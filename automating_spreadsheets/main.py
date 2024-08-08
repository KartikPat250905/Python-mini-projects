# This script processes an Excel file named 'transactions.xlsx' by performing the following operations:
# 1. Opens the workbook and accesses the first sheet.
# 2. Iterates over a range of rows to apply a 10% discount to the values in the third column.
# 3. Creates a bar chart based on the discounted values in the third column.
# 4. Adds the bar chart to the sheet at a specified location.
# 5. Saves the modified workbook as 'transactions2.xlsx'.

import openpyxl as xl
from openpyxl.chart import BarChart, Reference

wb = xl.load_workbook('transactions.xlsx')
sheet = wb['Sheet1']

for row in range(2, sheet.max_row + 1):
	cell = sheet.cell(row, 3)
	cell.value = cell.value * 0.9

values = Reference(sheet,min_row=2,max_row=4,min_col=3,max_col=3)
chart = BarChart()
chart.add_data(values)
sheet.add_chart(chart,'d2')
wb.save('transactions2.xlsx')
